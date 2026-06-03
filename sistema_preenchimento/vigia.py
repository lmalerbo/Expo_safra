"""
vigia.py — Monitora consolidar_queue/ e processa pedidos de consolidação.

Roda no PC servidor (24h). Não precisa de firewall nem de rede.
Comunicação via arquivos JSON no G: (drive compartilhado).

Uso: python vigia.py  (ou via INICIAR_VIGIA.bat)
"""

import json
import os
import subprocess
import sys
import time
from datetime import datetime

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERRO: openpyxl não encontrado. Execute: pip install openpyxl")
    input("Pressione Enter para sair...")
    sys.exit(1)

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
QUEUE_DIR = os.path.join(BASE_DIR, 'consolidar_queue')
EXP_DIR   = os.path.join(BASE_DIR, 'exports')
LOG_FILE  = os.path.join(BASE_DIR, 'logs', 'vigia.log')

# ── Carrega configurações ─────────────────────────────────────────────────
_cfg_path = os.path.join(BASE_DIR, 'config.json')
try:
    with open(_cfg_path, 'r', encoding='utf-8') as _f:
        _cfg = json.load(_f)
except Exception:
    _cfg = {}

INTERVALO         = _cfg.get('intervalo_vigia_s', 2)
TIMEOUT_CONSOLIDAR = _cfg.get('timeout_consolidar_s', 180)

COLS       = ['TIMESTAMP','USUARIO','LAYER','COD_FAZ','FAZENDA','TALHAO',
              'FRENTE','SEMANA','AREA_HA','TIPO_LINHA','CICLO','STATUS']
COL_WIDTHS = [18, 12, 12, 10, 28, 8, 8, 8, 9, 18, 8, 14]


# ── Logging (console + arquivo) ───────────────────────────────────────────
os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)

def _log(msg):
    line = f'[{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}] {msg}'
    print(line)
    try:
        with open(LOG_FILE, 'a', encoding='utf-8') as _lf:
            _lf.write(line + '\n')
    except Exception:
        pass


def _salvar_excel(records, usuario, dia, req_id):
    """Salva os registros em XLSX dentro de exports/.

    O req_id é incluído no nome para garantir unicidade mesmo quando o mesmo
    usuário envia múltiplas requisições no mesmo dia.
    """
    os.makedirs(EXP_DIR, exist_ok=True)
    fname = f'export_frentes_{usuario}_{dia}_{req_id}.xlsx'
    fpath = os.path.join(EXP_DIR, fname)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registros'
    ws.append(COLS)
    for r in records:
        ws.append([
            r.get('timestamp'), r.get('usuario'), r.get('layer'),
            r.get('cod_faz'),   r.get('fazenda'),  r.get('talhao'),
            r.get('frente'),    r.get('semana'),   r.get('ha', 0),
            r.get('tipo'),      r.get('ciclo'),    r.get('status'),
        ])
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(fpath)
    return fname


def _rodar_consolidar():
    script = os.path.join(BASE_DIR, 'engine', 'consolidar.py')
    try:
        env = os.environ.copy()
        env['PYTHONIOENCODING'] = 'utf-8'
        result = subprocess.run(
            [sys.executable, script],
            cwd=BASE_DIR,
            input='2\n\n',
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            timeout=TIMEOUT_CONSOLIDAR,
            env=env,
        )
        output = (result.stdout + result.stderr).strip()
        return result.returncode == 0, output
    except subprocess.TimeoutExpired:
        return False, f'ERRO: consolidar.py excedeu o tempo limite ({TIMEOUT_CONSOLIDAR}s).'
    except Exception as e:
        return False, f'ERRO ao executar consolidar.py: {e}'


def _processar(req_path):
    try:
        with open(req_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        _log(f'Erro ao ler {req_path}: {e}')
        return

    req_id  = data.get('id', 'unknown')
    records = data.get('records', [])
    usuario = data.get('usuario', 'usuario')
    dia     = data.get('dia', datetime.now().strftime('%Y-%m-%d'))
    res_path = os.path.join(QUEUE_DIR, f'res_{req_id}.json')

    _log(f'Processando pedido {req_id} — {len(records)} registros de {usuario}')

    try:
        arquivo = _salvar_excel(records, usuario, dia, req_id)
        _log(f'Excel salvo: {arquivo}')
    except Exception as e:
        _escrever_resposta(res_path, {'ok': False, 'arquivo': '', 'output': f'Erro ao salvar Excel: {e}'})
        _remover(req_path)
        return

    ok, output = _rodar_consolidar()
    status = 'OK' if ok else 'ERRO'
    _log(f'Consolidação {status} — {req_id}')

    _escrever_resposta(res_path, {'ok': ok, 'arquivo': arquivo, 'output': output})
    _remover(req_path)


def _escrever_resposta(path, data):
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception as e:
        _log(f'Erro ao escrever resposta: {e}')


def _remover(path):
    try:
        os.remove(path)
    except Exception:
        pass


_processados = set()  # IDs já tratados nesta sessão

def _verificar_queue():
    try:
        entries = os.listdir(QUEUE_DIR)
    except Exception:
        return
    for fname in entries:
        if not (fname.startswith('req_') and fname.endswith('.json')):
            continue
        req_id = fname[4:-5]
        if req_id in _processados:
            continue
        _processados.add(req_id)
        _processar(os.path.join(QUEUE_DIR, fname))


if __name__ == '__main__':
    os.makedirs(QUEUE_DIR, exist_ok=True)
    print('=' * 55)
    print('  Vigia PF — Monitorando consolidar_queue/')
    print(f'  Queue : {QUEUE_DIR}')
    print(f'  Base  : {BASE_DIR}')
    print(f'  Intervalo: {INTERVALO}s')
    print(f'  Log   : {LOG_FILE}')
    print('  Pressione Ctrl+C para parar.')
    print('=' * 55)
    _log('Vigia iniciado.')
    try:
        while True:
            _verificar_queue()
            time.sleep(INTERVALO)
    except KeyboardInterrupt:
        _log('Vigia encerrado pelo usuário.')
        print('\nVigia encerrado.')
