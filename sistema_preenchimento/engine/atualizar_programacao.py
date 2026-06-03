"""
atualizar_programacao.py
Uso: python atualizar_programacao.py  OU  duplo clique no ATUALIZAR.bat

Estrutura esperada:
  base_icol/*.xlsm          ← base ICOL (agendamento por frente)
  base_fazendas/*.xlsx      ← base mestre de talhões (área, estágio)
  programacao_frentes.xlsx
  formulario.html
"""

import os
import sys

# ── Utilitários compartilhados ────────────────────────────────────────────
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR   = os.path.dirname(_SCRIPT_DIR)   # sistema_preenchimento/
sys.path.insert(0, _SCRIPT_DIR)
from utils import layer_to_str, atualizar_html, redirecionar_stdout, fechar_log

_log_fh = redirecionar_stdout(os.path.join(_BASE_DIR, 'logs', 'atualizar.log'))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import datetime, time, json
import glob as _glob

# ── Carrega configurações ─────────────────────────────────────────────────
_cfg_path = os.path.join(_BASE_DIR, 'config.json')
try:
    with open(_cfg_path, 'r', encoding='utf-8') as _f:
        _cfg = json.load(_f)
except Exception:
    _cfg = {}

_col = _cfg.get('col_indices', {})
COL_LAYER   = _col.get('LAYER',      1)
COL_FRENTE  = _col.get('FRENTE',     2)
COL_PERIODO = _col.get('PERIODO_OP', 3)
COL_DIA     = _col.get('DIA',        4)
COL_CODFAZ  = _col.get('COD_FAZ',    5)
COL_FAZENDA = _col.get('FAZENDA',    6)
COL_TALHAO  = _col.get('TALHAO',     7)
COL_STATUS  = _col.get('STATUS',     8)   # cabeçalho na planilha: EXPORTAÇÃO
COL_TIPO    = _col.get('TIPO_LINHA', 9)   # cabeçalho: TIPO DE LINHA
COL_CICLO   = _col.get('CICLO',      10)
COL_AREA_HA = _col.get('AREA_HA',    11)
COL_ESTAGIO = _col.get('ESTAGIO',    12)

TENTATIVAS_BLOQUEIO     = _cfg.get('tentativas_arquivo_bloqueado', 3)
ESPERA_BLOQUEIO_S       = _cfg.get('espera_arquivo_bloqueado_s', 10)
CODFAZ_EXCLUIR_PREFIXO  = _cfg.get('codfaz_excluir_prefixo', '20')

DEST_FILE = "programacao_frentes.xlsx"
HTML_FILE = "formulario.html"

# ── Localiza o .xlsm em base_icol/ ───────────────────────────────────────
_xlsm_found = _glob.glob("base_icol/*.xlsm")
if not _xlsm_found:
    print("ERRO: Nenhum arquivo .xlsm encontrado em base_icol/")
    print("  Coloque a base ICOL na pasta base_icol/ e tente novamente.")
    fechar_log(_log_fh)
    input("\nPressione Enter para sair...")
    sys.exit(1)
SOURCE_XLSM = _xlsm_found[0]
print(f"Base ICOL encontrada: {SOURCE_XLSM}")

# ── Verificações ──────────────────────────────────────────────────────────
for f in [DEST_FILE, HTML_FILE]:
    if not os.path.exists(f):
        print(f"ERRO: Arquivo nao encontrado → {f}")
        fechar_log(_log_fh)
        input("\nPressione Enter para sair...")
        sys.exit(1)

def arquivo_bloqueado(path):
    try:
        with open(path, 'a+b'): return False
    except (IOError, PermissionError): return True

print("Verificando se a planilha esta aberta...")
tentativas = 0
while arquivo_bloqueado(DEST_FILE):
    tentativas += 1
    if tentativas >= TENTATIVAS_BLOQUEIO:
        print(f"ERRO: '{DEST_FILE}' está aberto. Feche e tente novamente.")
        fechar_log(_log_fh)
        input("\nPressione Enter para sair...")
        sys.exit(1)
    print(f"  Arquivo em uso, aguardando {ESPERA_BLOQUEIO_S}s... ({tentativas}/{TENTATIVAS_BLOQUEIO})")
    time.sleep(ESPERA_BLOQUEIO_S)
print("  OK.\n")

# ── 1. Ler valores existentes ─────────────────────────────────────────────
print("Lendo dados existentes na planilha...")
wb_ex = load_workbook(DEST_FILE, data_only=True)
ws_ex = wb_ex["Programação"]
preserved = {}   # layer_str → (exp, tipo, ciclo)
for row in ws_ex.iter_rows(min_row=2, values_only=True):
    layer = layer_to_str(row[0])
    if layer:
        preserved[layer] = (row[COL_STATUS-1] or '', row[COL_TIPO-1] or '', row[COL_CICLO-1] or '')
wb_ex.close()
print(f"  {len(preserved)} linhas existentes carregadas.\n")

# ── 2. Ler base_fazendas (mestre de talhões) ─────────────────────────────
print("Verificando base fazendas...")
_base_faz_files = _glob.glob("base_fazendas/*.xls*")
df_base = None
if not _base_faz_files:
    print("  AVISO: Nenhum arquivo em base_fazendas/ — usando apenas ICOL.\n")
else:
    SOURCE_BASE = _base_faz_files[0]
    print(f"  Base fazendas: {SOURCE_BASE}")
    df_base = pd.read_excel(SOURCE_BASE, engine='openpyxl')
    df_base = df_base.rename(columns={
        'SECAO':     'COD FAZ',
        'DESC_SECAO':'FAZENDA',
        'TALHAO':    'TALHOES',
        'AREA_PROD': 'AREA_HA',
    })
    df_base['COD FAZ'] = pd.to_numeric(df_base['COD FAZ'], errors='coerce')
    df_base['TALHOES'] = pd.to_numeric(df_base['TALHOES'], errors='coerce')
    df_base['AREA_HA'] = pd.to_numeric(df_base['AREA_HA'], errors='coerce')
    if 'ESTAGIO' not in df_base.columns:
        df_base['ESTAGIO'] = ''
    df_base = df_base.dropna(subset=['COD FAZ', 'TALHOES']).reset_index(drop=True)
    print(f"  {len(df_base)} talhões na base fazendas.\n")

# ── 3. Ler base ICOL (agendamento por fazenda) ───────────────────────────
print("Lendo base ICOL...")
df_raw = pd.read_excel(SOURCE_XLSM, sheet_name='BASE PARA PLANEJAMENTO', header=None, engine='openpyxl')

# Encontra primeira linha de dados (COD FAZ em AA=índice 26 deve ser numérico)
first_data_row = None
for idx, row_vals in df_raw.iterrows():
    try:
        val = float(str(row_vals.iloc[26]))
        if not pd.isna(val):
            first_data_row = idx
            break
    except (ValueError, TypeError):
        continue

if first_data_row is None:
    print("ERRO: Dados não encontrados na aba BASE PARA PLANEJAMENTO")
    print("  Verifique se a aba existe e se COD FAZENDA está na coluna AA.")
    fechar_log(_log_fh)
    input("\nPressione Enter para sair...")
    sys.exit(1)

# H=7 FRENTE, I=8 PERÍODO OP (mês), AA=26 COD FAZ, AB=27 FAZENDA
df_data = df_raw.iloc[first_data_row:, [7, 8, 26, 27]].copy()
df_data.columns = ['FRENTE', 'PERIODO_OP', 'COD FAZ', 'FAZENDA']

df_icol = df_data.dropna(subset=['COD FAZ']).copy()
df_icol['COD FAZ']    = pd.to_numeric(df_icol['COD FAZ'],    errors='coerce')
df_icol['FRENTE']     = pd.to_numeric(df_icol['FRENTE'],     errors='coerce')
df_icol['PERIODO_OP'] = pd.to_numeric(df_icol['PERIODO_OP'], errors='coerce')
df_icol = df_icol.dropna(subset=['COD FAZ'])
n_icol_raw = len(df_icol)
df_icol = df_icol.drop_duplicates(subset=['COD FAZ'], keep='first').reset_index(drop=True)
n_dup_icol = n_icol_raw - len(df_icol)
print(f"  {len(df_icol)} fazendas únicas no ICOL{f' ({n_dup_icol} linhas duplicadas ignoradas)' if n_dup_icol else ''}.\n")

# ── 4. Merge: ICOL (fazenda) LEFT JOIN base_fazendas (talhões) ───────────
if df_base is not None:
    result = df_icol[['COD FAZ', 'FAZENDA', 'FRENTE', 'PERIODO_OP']].merge(
        df_base[['COD FAZ', 'TALHOES', 'AREA_HA', 'ESTAGIO']],
        on='COD FAZ',
        how='left'
    )
    sem_ha = result['AREA_HA'].isna().sum()
    print(f"  {len(result)} talhões expandidos do ICOL — {len(result)-sem_ha} com AREA_HA, {sem_ha} sem.\n")
    faz_sem_base = result[result['TALHOES'].isna()].drop_duplicates('COD FAZ')
    n_sem_base = len(faz_sem_base)
    if n_sem_base:
        print(f"  ⚠  {n_sem_base} fazenda(s) do ICOL sem talhões na base_fazendas (serão omitidas do dashboard):")
        for _, r in faz_sem_base.head(10).iterrows():
            print(f"     COD FAZ {int(r['COD FAZ'])}: {r['FAZENDA']}")
        if n_sem_base > 10:
            print(f"     ... e mais {n_sem_base - 10}")
        print()
else:
    n_sem_base = 0
    result = df_icol.copy()
    result['TALHOES'] = None
    result['AREA_HA'] = None
    result['ESTAGIO'] = ''

result = result.sort_values(['FRENTE', 'PERIODO_OP']).reset_index(drop=True)

# Fazendas com COD FAZ iniciando no prefixo configurado são unidades administrativas
n_antes = result['COD FAZ'].nunique()
result   = result[~result['COD FAZ'].astype(str).str.startswith(CODFAZ_EXCLUIR_PREFIXO)].reset_index(drop=True)
n_excluidas = n_antes - result['COD FAZ'].nunique()
if n_excluidas:
    print(f"  Filtro administrativo (COD FAZ {CODFAZ_EXCLUIR_PREFIXO}x): {n_excluidas} fazenda(s) excluída(s).\n")

def make_layer(row):
    try: return int(f"{int(row['COD FAZ'])}{int(row['TALHOES']):03d}")
    except (ValueError, TypeError, KeyError): return None

result['LAYER'] = result.apply(make_layer, axis=1)

# ha_lookup: (COD FAZ, TALHOES) → AREA_HA
ha_lookup = {}
if df_base is not None:
    for _, row in df_base.iterrows():
        try:
            key = (int(row['COD FAZ']), int(row['TALHOES']))
            if pd.notna(row.get('AREA_HA')):
                ha_lookup[key] = round(float(row['AREA_HA']), 2)
        except (ValueError, TypeError):
            pass

# layer_ha: layer_str → AREA_HA (para cálculo de ha por usuário)
layer_ha = {}
for _, row in result.iterrows():
    ly = layer_to_str(row.get('LAYER'))
    if ly:
        try:
            layer_ha[ly] = ha_lookup.get((int(row['COD FAZ']), int(row['TALHOES'])), 0)
        except (ValueError, TypeError):
            pass

print(f"  {len(result)} linhas para processar.\n")

# ── 5. Atualizar planilha ─────────────────────────────────────────────────
print("Atualizando planilha...")
wb = load_workbook(DEST_FILE)
ws = wb["Programação"]
ws.delete_rows(2, ws.max_row)

frente_row_colors = {2:"DDEEFF",3:"E8F2FB",4:"D6E4F5",5:"EBF3FF",
                     6:"F2F8FD",8:"DDEEFF",9:"E8F2FB",10:"D6E4F5"}
thin   = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Garante cabeçalhos para colunas que podem não existir em templates antigos
if ws.cell(1, COL_PERIODO).value not in ('PERÍODO OP', 'PERÍODO OPERACIONAL'):
    h = ws.cell(1, COL_PERIODO, value='PERÍODO OP')
    h.font      = Font(name="Arial", bold=True, size=10)
    h.border    = border
    h.alignment = Alignment(horizontal="center", vertical="center")
if ws.cell(1, COL_AREA_HA).value != 'AREA_HA':
    h = ws.cell(1, COL_AREA_HA, value='AREA_HA')
    h.font      = Font(name="Arial", bold=True, size=10)
    h.border    = border
    h.alignment = Alignment(horizontal="center", vertical="center")
if ws.cell(1, COL_ESTAGIO).value != 'ESTÁGIO':
    h = ws.cell(1, COL_ESTAGIO, value='ESTÁGIO')
    h.font      = Font(name="Arial", bold=True, size=10)
    h.border    = border
    h.alignment = Alignment(horizontal="center", vertical="center")

novos = 0
for i, row in result.iterrows():
    r      = i + 2
    frente = int(row['FRENTE']) if pd.notna(row.get('FRENTE')) else 0
    cor    = frente_row_colors.get(frente, "FFFFFF")
    fill   = PatternFill("solid", start_color=cor, end_color=cor)
    ly_str = layer_to_str(row.get('LAYER'))
    if ly_str in preserved:
        exp, tipo, ciclo = preserved[ly_str]
    else:
        exp, tipo, ciclo = '', '', ''
        novos += 1
    try:
        ha_val = ha_lookup.get((int(row['COD FAZ']), int(row['TALHOES'])), 0)
    except (ValueError, TypeError):
        ha_val = 0
    estagio = str(row.get('ESTAGIO', '') or '').strip()
    periodo = int(row['PERIODO_OP']) if pd.notna(row.get('PERIODO_OP')) else ''
    layer_val = row['LAYER']
    vals = [
        layer_val if layer_val else '', frente,
        periodo,
        '',
        int(row['COD FAZ']) if pd.notna(row['COD FAZ']) else '',
        str(row['FAZENDA']) if pd.notna(row['FAZENDA']) else '',
        int(row['TALHOES']) if pd.notna(row.get('TALHOES')) else '',
        exp, tipo, ciclo,
        ha_val,
        estagio,
    ]
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font      = Font(name="Arial", size=10)
        cell.fill      = fill
        cell.border    = border
        cell.alignment = Alignment(horizontal="left" if c==COL_FAZENDA else "center", vertical="center")
        if c == COL_AREA_HA:
            cell.number_format = '#,##0.00'

wb.save(DEST_FILE)
print(f"  Planilha atualizada. Novas linhas: {novos}\n")

# ── Aviso: LAYERs com preenchimento removidos desta atualização ───────────
layers_novos_str = {layer_to_str(ly) for ly in result['LAYER'].dropna() if layer_to_str(str(ly))}
layers_com_preenchimento = {ly for ly, vals in preserved.items() if any(str(v).strip() for v in vals)}
removidos_icol = layers_com_preenchimento - layers_novos_str
if removidos_icol:
    print(f"  ⚠  ATENÇÃO: {len(removidos_icol)} LAYER(s) preenchidos foram removidos desta base ICOL:")
    for ly in sorted(removidos_icol)[:10]:
        exp, tipo, ciclo = preserved[ly]
        print(f"     LAYER {ly}: {exp} | {tipo} | {ciclo}")
    if len(removidos_icol) > 10:
        print(f"     ... e mais {len(removidos_icol)-10}")
    print("     Esses dados NÃO foram perdidos — apenas saíram do ICOL atual.\n")

# ── 6. Gerar dados para o HTML ────────────────────────────────────────────
print("Gerando dados para o formulario...")
df_html = pd.read_excel(DEST_FILE)

fazendas_data = {}
for _, row in df_html.iterrows():
    faz  = str(row['FAZENDA']).strip()
    tal  = int(row['TALHÕES']) if pd.notna(row['TALHÕES']) else None
    layer = int(row['LAYER']) if pd.notna(row['LAYER']) else None
    cod  = int(row['COD FAZ']) if pd.notna(row['COD FAZ']) else None
    frente = int(row['FRENTE']) if pd.notna(row['FRENTE']) else None
    semana = int(row['PERÍODO OP']) if pd.notna(row.get('PERÍODO OP')) else None
    exp  = str(row['EXPORTAÇÃO']).strip() if pd.notna(row['EXPORTAÇÃO']) and str(row['EXPORTAÇÃO']) not in ['','nan'] else ''
    tipo = str(row['TIPO DE LINHA']).strip() if pd.notna(row['TIPO DE LINHA']) and str(row['TIPO DE LINHA']) not in ['','nan'] else ''
    ciclo = str(row['CICLO']).strip() if pd.notna(row['CICLO']) and str(row['CICLO']) not in ['','nan'] else ''

    if not faz or faz == 'nan' or tal is None: continue

    if faz not in fazendas_data:
        fazendas_data[faz] = {'cod': cod, 'talhoes': {}}

    tal_key = str(tal)
    if tal_key not in fazendas_data[faz]['talhoes']:
        ha = ha_lookup.get((cod, tal), 0)
        estagio_str = str(row.get('ESTÁGIO', '') or '').strip()
        fazendas_data[faz]['talhoes'][tal_key] = {
            'layer': layer, 'frente': frente, 'semana': semana,
            'exp': exp, 'tipo': tipo, 'ciclo': ciclo, 'ha': ha, 'dia': None,
            'estagio': estagio_str
        }

n_html_faz = len(fazendas_data)
n_html_tal = sum(len(v['talhoes']) for v in fazendas_data.values())
print(f"  {n_html_faz} fazendas, {n_html_tal} talhões → HTML.\n")
print(f"  Resumo do funil de fazendas:")
print(f"    ICOL bruto                : {n_icol_raw}")
if n_dup_icol:      print(f"    (-) Duplicatas COD FAZ    : {n_dup_icol}")
if n_excluidas:     print(f"    (-) Administrativas ({CODFAZ_EXCLUIR_PREFIXO}x) : {n_excluidas}")
if n_sem_base:      print(f"    (-) Sem talhão na base    : {n_sem_base}")
print(f"    (=) No dashboard          : {n_html_faz}\n")

# Ler Log Exportações → ha por usuário
print("Lendo log de exportacoes...")
usuario_ha = {}
LOG_SHEET = "Log Exportações"
try:
    df_log = pd.read_excel(DEST_FILE, sheet_name=LOG_SHEET)
    df_log_dedup = df_log.drop_duplicates(subset='LAYER', keep='last')
    for _, row in df_log_dedup.iterrows():
        usuario = str(row.get('USUARIO', '') or '').strip()
        if not usuario or usuario in ('nan', ''): continue
        ly = layer_to_str(row.get('LAYER', ''))
        if ly:
            usuario_ha[usuario] = round(usuario_ha.get(usuario, 0) + layer_ha.get(ly, 0), 2)
    print(f"  {len(usuario_ha)} usuário(s) no log.\n")
except Exception as e:
    print(f"  AVISO: log nao encontrado ou vazio ({e})\n")

# Ler configuração de usuários
usuarios_config = []
if os.path.exists('usuarios.json'):
    try:
        with open('usuarios.json', 'r', encoding='utf-8') as f:
            usuarios_config = json.load(f)
    except Exception as e:
        print(f"  AVISO: nao foi possivel ler usuarios.json ({e})\n")

# ── 7. Atualizar HTML ─────────────────────────────────────────────────────
print("Atualizando formulario.html...")
ok, msg = atualizar_html(HTML_FILE, fazendas_data, usuario_ha, usuarios_config)
if ok:
    print(f"  formulario.html atualizado — {msg}.\n")
else:
    print(f"  AVISO: {msg}\n")

# ── Resumo ────────────────────────────────────────────────────────────────
print(f"{'='*50}")
print(f"  Atualizacao concluida!")
print(f"  Total talhoes : {len(result)}")
if df_base is not None:
    sem_ha = result['AREA_HA'].isna().sum()
    print(f"  Com AREA_HA   : {len(result) - sem_ha}")
    print(f"  Sem AREA_HA   : {sem_ha}")
print(f"  Preservados   : {len(result) - novos}")
print(f"  HTML regerado : {'sim' if ok else 'FALHOU — ' + msg}")
print(f"{'='*50}")

fechar_log(_log_fh)
input("\nPressione Enter para fechar...")
