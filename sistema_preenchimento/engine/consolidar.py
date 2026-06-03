"""
consolidar.py — Consolida os exports diários na planilha mestre

Uso:
  1. Coloque este script na mesma pasta que programacao_frentes.xlsm
  2. Crie uma subpasta chamada  exports/
  3. Mova todos os arquivos export_frentes_*.xlsx para essa pasta
  4. Execute: python consolidar.py

O script:
  - Lê todos os exports da pasta exports/
  - Para cada registro, sobrescreve as colunas EXPORTAÇÃO, TIPO DE LINHA, CICLO
    na linha correspondente ao LAYER na planilha mestre
  - Move os arquivos processados para exports/processados/
  - Gera um relatório resumido no terminal
"""

import os
import sys

# ── Utilitários compartilhados ────────────────────────────────────────────
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR   = os.path.dirname(_SCRIPT_DIR)   # sistema_preenchimento/
sys.path.insert(0, _SCRIPT_DIR)
from utils import layer_to_str, atualizar_html, redirecionar_stdout, fechar_log

_log_fh = redirecionar_stdout(os.path.join(_BASE_DIR, 'logs', 'consolidar.log'))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import datetime
import shutil
import glob
import json
import time

# ── Carrega configurações ─────────────────────────────────────────────────
_cfg_path = os.path.join(_BASE_DIR, 'config.json')
try:
    with open(_cfg_path, 'r', encoding='utf-8') as _f:
        _cfg = json.load(_f)
except Exception:
    _cfg = {}

_col = _cfg.get('col_indices', {})
COL_LAYER   = _col.get('LAYER',      1)
COL_FRENTE  = _col.get('FRENTE',     2)
COL_PERIODO = _col.get('PERIODO_OP', 3)
COL_DIA     = _col.get('DIA',        4)
COL_CODFAZ  = _col.get('COD_FAZ',    5)
COL_FAZENDA = _col.get('FAZENDA',    6)
COL_TALHAO  = _col.get('TALHAO',     7)
COL_STATUS  = _col.get('STATUS',     8)   # cabeçalho na planilha: EXPORTAÇÃO
COL_TIPO    = _col.get('TIPO_LINHA', 9)   # cabeçalho: TIPO DE LINHA
COL_CICLO   = _col.get('CICLO',      10)
COL_AREA_HA = _col.get('AREA_HA',    11)
COL_ESTAGIO = _col.get('ESTAGIO',    12)

TENTATIVAS_BLOQUEIO = _cfg.get('tentativas_arquivo_bloqueado', 3)
ESPERA_BLOQUEIO_S   = _cfg.get('espera_arquivo_bloqueado_s', 10)

MESTRE    = "programacao_frentes.xlsx"
PASTA_EXP = "exports"
PASTA_OK  = os.path.join(PASTA_EXP, "processados")

# Colunas obrigatórias nos arquivos de export gerados pelo formulario.html
EXPORT_COLS_REQUIRED = {'LAYER', 'STATUS', 'TIPO_LINHA', 'CICLO', 'USUARIO', 'TIMESTAMP'}

# ── Verificações iniciais ─────────────────────────────────────────────────
for f in [MESTRE]:
    if not os.path.exists(f):
        print(f"ERRO: Arquivo nao encontrado → {f}")
        fechar_log(_log_fh)
        input("\nPressione Enter para sair...")
        sys.exit(1)

os.makedirs(PASTA_OK, exist_ok=True)

# ── Verifica se mestre está aberto ────────────────────────────────────────
def arquivo_bloqueado(path):
    try:
        with open(path, 'a+b'): return False
    except (IOError, PermissionError): return True

tentativas = 0
while arquivo_bloqueado(MESTRE):
    tentativas += 1
    if tentativas >= TENTATIVAS_BLOQUEIO:
        print(f"ERRO: '{MESTRE}' está aberto. Feche o arquivo e tente novamente.")
        fechar_log(_log_fh)
        input("\nPressione Enter para sair...")
        sys.exit(1)
    print(f"  Arquivo em uso, aguardando {ESPERA_BLOQUEIO_S}s... ({tentativas}/{TENTATIVAS_BLOQUEIO})")
    time.sleep(ESPERA_BLOQUEIO_S)

# ── Carrega exports ───────────────────────────────────────────────────────
exports = glob.glob(os.path.join(PASTA_EXP, "export_frentes_*.xlsx"))
if not exports:
    print(f"Nenhum arquivo export_frentes_*.xlsx encontrado em '{PASTA_EXP}/'")
    fechar_log(_log_fh)
    input("\nPressione Enter para sair...")
    sys.exit(0)

print(f"Encontrados {len(exports)} arquivo(s) de export:\n")
all_records = []
for path in exports:
    try:
        df = pd.read_excel(path)
        df['_source'] = os.path.basename(path)
        all_records.append(df)
        print(f"  ✓ {os.path.basename(path)} — {len(df)} registros")
    except Exception as e:
        print(f"  ✗ {os.path.basename(path)} — ERRO: {e}")

if not all_records:
    print("\nNenhum registro válido encontrado.")
    fechar_log(_log_fh)
    input("\nPressione Enter para sair...")
    sys.exit(0)

df_all = pd.concat(all_records, ignore_index=True)
print(f"\nTotal de registros para processar: {len(df_all)}")

# Normaliza LAYER para string inteira padronizada via layer_to_str
df_all['LAYER'] = df_all['LAYER'].apply(layer_to_str)

# Valida colunas obrigatórias antes de prosseguir
missing_cols = EXPORT_COLS_REQUIRED - set(df_all.columns)
if missing_cols:
    print(f"\nERRO: Colunas ausentes nos exports: {', '.join(sorted(missing_cols))}")
    print("  Verifique se os exports foram gerados pelo formulario.html correto.")
    fechar_log(_log_fh)
    input("\nPressione Enter para sair...")
    sys.exit(1)

df_dedup = df_all.drop_duplicates(subset='LAYER', keep='last')
print(f"Após deduplicação por LAYER: {len(df_dedup)} registros únicos")

# Monta lookup layer → valores (usa df_dedup: apenas 1 registro por LAYER)
lookup = df_dedup.set_index('LAYER')[['STATUS','TIPO_LINHA','CICLO','USUARIO','TIMESTAMP']].to_dict('index')

# ── Verifica conflitos (layers que já têm preenchimento na mestre) ─────────
print("\nVerificando conflitos na planilha mestre...")
wb = load_workbook(MESTRE)
ws = wb["Programação"]
lastRow = ws.max_row

conflitos = []   # (row, layer, val_mestre, val_export)
limpos    = []   # (row, layer) sem preenchimento atual

for row in range(2, lastRow + 1):
    layer_str = layer_to_str(ws.cell(row=row, column=COL_LAYER).value)
    if not layer_str or layer_str not in lookup:
        continue

    status_atual = ws.cell(row=row, column=COL_STATUS).value or ''
    tipo_atual   = ws.cell(row=row, column=COL_TIPO).value or ''
    ciclo_atual  = ws.cell(row=row, column=COL_CICLO).value or ''

    ja_preenchido = any([str(status_atual).strip(), str(tipo_atual).strip(), str(ciclo_atual).strip()])
    rec = lookup[layer_str]

    if ja_preenchido:
        conflitos.append({
            'row': row, 'layer': layer_str,
            'mestre': f"{status_atual} | {tipo_atual} | {ciclo_atual}",
            'export': f"{rec['STATUS']} | {rec['TIPO_LINHA']} | {rec['CICLO']}",
            'usuario': rec['USUARIO'], 'ts': rec['TIMESTAMP']
        })
    else:
        limpos.append({'row': row, 'layer': layer_str})

print(f"  Sem conflito (serão preenchidos): {len(limpos)}")
print(f"  Com conflito (já preenchidos)   : {len(conflitos)}")

# ── Resolve conflitos interativamente ─────────────────────────────────────
sobrescrever_ids = set()   # layers aprovados para sobrescrever

if conflitos:
    print(f"\n{'='*60}")
    print(f"  ATENÇÃO: {len(conflitos)} talhão(ões) já têm preenchimento.")
    print(f"{'='*60}")
    print("\nO que deseja fazer?\n")
    print("  1 - Ver cada conflito e decidir um por um")
    print("  2 - Sobrescrever TODOS automaticamente")
    print("  3 - Manter TODOS (não sobrescreve nenhum conflito)")
    print()
    escolha = input("Digite 1, 2 ou 3: ").strip()

    if escolha == '2':
        sobrescrever_ids = {c['layer'] for c in conflitos}
        print(f"\n  → Todos os {len(conflitos)} conflitos serão sobrescritos.")

    elif escolha == '1':
        print()
        for i, c in enumerate(conflitos, 1):
            print(f"  [{i}/{len(conflitos)}] LAYER {c['layer']}")
            print(f"    Mestre : {c['mestre']}")
            print(f"    Export : {c['export']}  (por {c['usuario']} em {c['ts']})")
            resp = input("    Sobrescrever? (s/N): ").strip().lower()
            if resp == 's':
                sobrescrever_ids.add(c['layer'])
        print(f"\n  → {len(sobrescrever_ids)} de {len(conflitos)} conflitos serão sobrescritos.")

    else:  # '3' ou qualquer outra coisa
        print(f"\n  → Nenhum conflito será sobrescrito. Apenas novos registros serão aplicados.")

# ── Aplica atualizações ───────────────────────────────────────────────────
print("\nAplicando atualizações...")
atualizados = 0
ignorados   = 0
nao_encontrados = []

for row in range(2, lastRow + 1):
    layer_str = layer_to_str(ws.cell(row=row, column=COL_LAYER).value)
    if not layer_str or layer_str not in lookup:
        continue

    is_conflito = any(c['layer'] == layer_str for c in conflitos)
    if is_conflito and layer_str not in sobrescrever_ids:
        ignorados += 1
        continue

    rec = lookup[layer_str]
    ws.cell(row=row, column=COL_STATUS).value = str(rec['STATUS'])
    ws.cell(row=row, column=COL_TIPO).value   = str(rec['TIPO_LINHA'])
    ws.cell(row=row, column=COL_CICLO).value  = str(rec['CICLO'])
    atualizados += 1

# Verifica layers não encontrados na mestre
layers_mestre = set()
for row in range(2, lastRow + 1):
    v = layer_to_str(ws.cell(row=row, column=COL_LAYER).value)
    if v:
        layers_mestre.add(v)

for layer in lookup:
    if layer not in layers_mestre:
        nao_encontrados.append(layer)

# ── Aba de log histórico ──────────────────────────────────────────────────
LOG_SHEET = "Log Exportações"
if LOG_SHEET not in wb.sheetnames:
    ws_log = wb.create_sheet(LOG_SHEET)
    headers = ['DATA_CONSOLIDACAO','TIMESTAMP','USUARIO','LAYER','FAZENDA','TALHAO','TIPO_LINHA','CICLO','STATUS','ARQUIVO']
    thin = Side(style="thin", color="BFBFBF")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    azul = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    for c, h in enumerate(headers, 1):
        cell = ws_log.cell(row=1, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = azul
        cell.border = brd
        cell.alignment = Alignment(horizontal="center")
else:
    ws_log = wb[LOG_SHEET]

last_log = ws_log.max_row
now_str  = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
for _, rec in df_all.iterrows():
    last_log += 1
    ws_log.cell(row=last_log, column=1).value  = now_str
    ws_log.cell(row=last_log, column=2).value  = str(rec.get('TIMESTAMP',''))
    ws_log.cell(row=last_log, column=3).value  = str(rec.get('USUARIO',''))
    ws_log.cell(row=last_log, column=4).value  = str(rec.get('LAYER',''))
    ws_log.cell(row=last_log, column=5).value  = str(rec.get('FAZENDA',''))
    ws_log.cell(row=last_log, column=6).value  = str(rec.get('TALHAO',''))
    ws_log.cell(row=last_log, column=7).value  = str(rec.get('TIPO_LINHA',''))
    ws_log.cell(row=last_log, column=8).value  = str(rec.get('CICLO',''))
    ws_log.cell(row=last_log, column=9).value  = str(rec.get('STATUS',''))
    ws_log.cell(row=last_log, column=10).value = str(rec.get('_source',''))
    for c in range(1, 11):
        ws_log.cell(row=last_log, column=c).font = Font(name="Arial", size=10)

wb.save(MESTRE)

# ── Regenerar formulario.html ─────────────────────────────────────────────
HTML_FILE = "formulario.html"
if os.path.exists(HTML_FILE):
    print("\nRegenerando formulario.html...")
    try:
        df_mestre = pd.read_excel(MESTRE)

        def _safe_int(v):
            try: return int(v) if pd.notna(v) else None
            except (ValueError, TypeError): return None

        def _safe_str(v):
            s = str(v or '').strip()
            return '' if s == 'nan' else s

        layer_ha = {}
        for _, row in df_mestre.iterrows():
            ly = layer_to_str(row.get('LAYER', ''))
            if ly:
                try: layer_ha[ly] = round(float(row.get('AREA_HA', 0) or 0), 2)
                except (ValueError, TypeError): pass

        fazendas_data = {}
        for _, row in df_mestre.iterrows():
            faz = _safe_str(row.get('FAZENDA', ''))
            tal = row.get('TALHÕES')
            if not faz or pd.isna(tal):
                continue
            try: tal = int(tal)
            except (ValueError, TypeError): continue

            if faz not in fazendas_data:
                fazendas_data[faz] = {'cod': _safe_int(row.get('COD FAZ')), 'talhoes': {}}

            tal_key = str(tal)
            if tal_key not in fazendas_data[faz]['talhoes']:
                fazendas_data[faz]['talhoes'][tal_key] = {
                    'layer':  _safe_int(row.get('LAYER')),
                    'frente': _safe_int(row.get('FRENTE')),
                    'semana': _safe_int(row.get('PERÍODO OP')),
                    'exp':    _safe_str(row.get('EXPORTAÇÃO', '')),
                    'tipo':   _safe_str(row.get('TIPO DE LINHA', '')),
                    'ciclo':  _safe_str(row.get('CICLO', '')),
                    'ha':     float(row.get('AREA_HA', 0) or 0),
                    'dia':    None,
                    'estagio': _safe_str(row.get('ESTÁGIO', '')),
                }

        usuario_ha = {}
        try:
            df_log = pd.read_excel(MESTRE, sheet_name=LOG_SHEET)
            df_log_dedup = df_log.drop_duplicates(subset='LAYER', keep='last')
            for _, row in df_log_dedup.iterrows():
                usuario = _safe_str(row.get('USUARIO', ''))
                if not usuario: continue
                ly = layer_to_str(row.get('LAYER', ''))
                if ly:
                    usuario_ha[usuario] = round(usuario_ha.get(usuario, 0) + layer_ha.get(ly, 0), 2)
        except Exception as e:
            print(f"  AVISO: nao foi possivel calcular USUARIOS_HA do log ({e})")

        usuarios_config = []
        if os.path.exists('usuarios.json'):
            try:
                with open('usuarios.json', 'r', encoding='utf-8') as _f:
                    usuarios_config = json.load(_f)
            except Exception as e:
                print(f"  AVISO: nao foi possivel ler usuarios.json ({e})")

        ok, msg = atualizar_html(HTML_FILE, fazendas_data, usuario_ha, usuarios_config)
        if ok:
            print(f"  formulario.html atualizado — {msg}.")
        else:
            print(f"  AVISO: {msg}")

    except Exception as e:
        print(f"  ERRO ao regenerar HTML: {e}")

# ── Move exports processados ──────────────────────────────────────────────
movidos = 0
for path in exports:
    if not os.path.exists(path):
        print(f"  AVISO: {os.path.basename(path)} não encontrado (já movido?). Ignorando.")
        continue
    dest = os.path.join(PASTA_OK, os.path.basename(path))
    try:
        if os.path.exists(dest):
            os.remove(dest)
        shutil.move(path, dest)
        movidos += 1
    except Exception as e:
        try:
            shutil.copy2(path, dest)
            os.remove(path)
            movidos += 1
        except Exception as e2:
            print(f"  AVISO: não foi possível mover {os.path.basename(path)}: {e2}")

# ── Relatório final ───────────────────────────────────────────────────────
print(f"\n{'='*50}")
print(f"  Consolidação concluída!")
print(f"  Registros processados : {len(df_all)}")
print(f"  Linhas atualizadas    : {atualizados}")
print(f"  Conflitos ignorados   : {ignorados}")
print(f"  Arquivos movidos para : {PASTA_OK}/ ({movidos}/{len(exports)})")
if nao_encontrados:
    print(f"\n  ⚠  {len(nao_encontrados)} LAYER(s) não encontrados na mestre:")
    for l in nao_encontrados[:10]:
        print(f"     - {l}")
    if len(nao_encontrados) > 10:
        print(f"     ... e mais {len(nao_encontrados)-10}")
print(f"  Log salvo na aba '{LOG_SHEET}'")
print(f"{'='*50}")

fechar_log(_log_fh)
input("\nPressione Enter para fechar...")
